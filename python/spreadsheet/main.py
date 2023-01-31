import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value  # Value for supplier
    inventory = product_list.cell(product_row, 2).value  # Value for inventory
    price = product_list.cell(product_row, 3).value  # Value for price
    product_num = product_list.cell(product_row, 1).value  # Value for product number
    inventory_price = product_list.cell(product_row, 5)  # Add column for total cost

    # Calculation for number of products per supplier
    # Adding to the product count each time we find a product for the same supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        # if it's a new supplier the count should equal 1
        # print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

    # Calculation for total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Product with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # Add value for total inventory price
    inventory_price.value = inventory * price

print("-" * 100)
print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)
print("-" * 100)

inv_file.save("inventory_with_total_value.xlsx")
